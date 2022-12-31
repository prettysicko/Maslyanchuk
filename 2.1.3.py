import csv
import math
import pathlib
import pdfkit

from matplotlib import pyplot as plt, figure
from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import worksheet
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from jinja2 import Environment, FileSystemLoader


class Salary:
    currency_to_rub = {"AZN": 35.68,
                       "BYR": 23.91,
                       "EUR": 59.90,
                       "GEL": 21.74,
                       "KGS": 0.76,
                       "KZT": 0.13,
                       "RUR": 1,
                       "UAH": 1.64,
                       "USD": 60.66,
                       "UZS": 0.0055}

    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def turn_ruble(self) -> float:
        return self.turn()

    def turn(self):
        value = float(self.currency_to_rub[self.salary_currency])
        return ((float(self.salary_from) + float(self.salary_to)) / 2) * value


class Vacancy:
    def __init__(self, vacancy_dict):
        self.table_first_line(vacancy_dict)

    def table_first_line(self, vacancy_dict):
        self.dict = vacancy_dict
        self.salary = Salary(self.dict['salary_from'],
                             self.dict['salary_to'],
                             self.dict['salary_currency'])


class Input:
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession = input("Введите название профессии: ")
        self.fields = []

    def parser_cod(self, all_vacancy, read):
        for row in read:
            new_vacancy = self.gain_vacancies(row)
            all_vacancy.append(new_vacancy)
        return all_vacancy

    def pars(self):
        read = self.read_cSV(self.file_name)
        all_vacancy = []
        return self.parser_cod(all_vacancy, read)

    def read_cSV(self, file_name) -> []:
        with open(file_name, encoding="utf-8-sig") as test:
            unpack = csv.reader(test)
            date = []
            c = 0
            for row in unpack:
                if c < len(row):
                    c = len(row)
                if '' not in row and c == len(row):
                    date.append(row)
            self.fields = date[0]
            return date[1:]

    def gain_vacancies(self, row: []) -> Vacancy:
        vacancy = Vacancy(dict(zip(self.fields, row)))
        return vacancy


class GraphData:
    def __init__(self, data, x_axis, profession="not"):
        self.data = data
        self.profession = profession
        self.salary_data = dict()
        self.count_data = dict()
        self.calc_date(x_axis)
        self.x_axis = x_axis

    def data_from_vacancies(self, vacancy: Vacancy, x_axis):
        if x_axis == "years":
            abs = int(vacancy.dict['published_at'].split('-')[0])
        else:
            abs = vacancy.dict['area_name']
        if abs not in self.salary_data:
            self.salary_data[abs] = 0
        if abs not in self.count_data:
            self.count_data[abs] = 0
        salary = vacancy.salary.turn_ruble()
        if self.profession != "not" and self.profession not in vacancy.dict['name']:
            return
        self.renew_direct(abs, salary)

    def calc_date(self, x_axis):
        for vacancy in self.data:
            self.data_from_vacancies(vacancy, x_axis)
        for x in self.salary_data:
            if self.count_data[x] != 0:
                self.salary_data[x] = math.floor(self.salary_data[x] / self.count_data[x])

    def renew_direct(self, key: str, value: float):
        self.renew(key, value)

    def renew(self, key, value):
        try:
            self.salary_data[key] += value
            self.count_data[key] += 1
        except:
            self.salary_data[key] = value
            self.count_data[key] = 1

    def gain_graph(self):
        first_printed_dict = self.salary_data
        second_printed_dict = self.count_data
        if self.x_axis == "areas":
            vac_count = sum(list(second_printed_dict.values()))
            first_printed_dict = self.sorted_dict(
                dict(list(filter(lambda x: self.count_data[x[0]] / vac_count > 0.01, self.salary_data.items()))))
            second_printed_dict = self.sorted_dict(
                dict(list(filter(lambda x: self.count_data[x[0]] / vac_count > 0.01, self.count_data.items()))))
            for x in second_printed_dict:
                second_printed_dict[x] = float("%.4f" % (second_printed_dict[x] / vac_count))
        return first_printed_dict, second_printed_dict

    @classmethod
    def sorted_dict(cls, non_sorted_dict: dict) -> dict:
        return dict(list(sorted(non_sorted_dict.items(), key=lambda x: x[1], reverse=True))[:10])


class ExcelReport:
    def __init__(self, side, font):
        self.border = Border(left=side, top=side, right=side, bottom=side)
        self.font = font

    def exel(self, direct: list):
        work_book = Workbook()
        years_sheet = self.make_years(work_book, direct[0], direct[1], direct[2], direct[3])
        areas_sheet = self.make_areas(work_book, direct[4], direct[5])
        ws = work_book.worksheets[0]
        work_book.remove(ws)
        work_book.save("report.xlsx")

    def make_years(self, work_book: Workbook, salaries: dict, prof_salaries: dict, counts: dict,
                   prof_counts: dict) -> worksheet.Worksheet:
        year_leaf = work_book.create_sheet("Статистика по годам")
        year_leaf['A1'] = 'Год'
        year_leaf['B1'] = 'Средняя зарплата'
        year_leaf['C1'] = f'Средняя зарплата - {input_set.profession}'
        year_leaf['D1'] = 'Количество вакансий'
        year_leaf['E1'] = f'Количество вакансий - {input_set.profession}'
        column_cells = [year_leaf["A1"], year_leaf['B1'], year_leaf['C1'], year_leaf['D1'], year_leaf['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        dim_holder = self.pays(counts, prof_counts, prof_salaries, salaries, year_leaf)

        for col in range(year_leaf.min_column, year_leaf.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(year_leaf, min=col, max=col, width=20)

        year_leaf.column_dimensions = dim_holder

        return year_leaf

    def pays(self, counts, prof_counts, prof_salaries, salaries, year_leaf):
        for x in range(len(salaries)):
            year_leaf[f'A{x + 2}'] = list(salaries.keys())[x]
            year_leaf[f'B{x + 2}'] = list(salaries.values())[x]
            year_leaf[f'C{x + 2}'] = list(prof_salaries.values())[x]
            year_leaf[f'D{x + 2}'] = list(counts.values())[x]
            year_leaf[f'E{x + 2}'] = list(prof_counts.values())[x]
            year_leaf[f'A{x + 2}'].border = self.border
            year_leaf[f'B{x + 2}'].border = self.border
            year_leaf[f'C{x + 2}'].border = self.border
            year_leaf[f'D{x + 2}'].border = self.border
            year_leaf[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=year_leaf)
        return dim_holder

    def make_areas(self, work_book: Workbook, salaries: dict, counts: dict) -> worksheet.Worksheet:
        area_leaf = work_book.create_sheet("Статистика по городам")
        area_leaf['A1'] = "Город"
        area_leaf['B1'] = "Уровень зарплат"
        area_leaf['D1'] = "Город"
        area_leaf['E1'] = "Доля вакансий"
        column_cells = [area_leaf['A1'], area_leaf['B1'], area_leaf['D1'], area_leaf['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        dim_holder = self.place(area_leaf, counts, salaries)

        for col in range(area_leaf.min_column, area_leaf.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(area_leaf, min=col, max=col, width=20)

        area_leaf.column_dimensions = dim_holder
        return area_leaf

    def place(self, area_leaf, counts, salaries):
        for x in range(10):
            area_leaf[f'A{x + 2}'] = list(salaries.keys())[x]
            area_leaf[f'B{x + 2}'] = list(salaries.values())[x]
            area_leaf[f'D{x + 2}'] = list(counts.keys())[x]
            area_leaf[f'E{x + 2}'] = list(counts.values())[x]
            area_leaf[f'A{x + 2}'].border = self.border
            area_leaf[f'B{x + 2}'].border = self.border
            area_leaf[f'D{x + 2}'].border = self.border
            area_leaf[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=area_leaf)
        return dim_holder


class PngReport:
    def __init__(self, years_salary: dict, years_count: dict,
                 prof_salary: dict, prof_count: dict,
                 areas_salary: dict, areas_count: dict):
        self.years_salary = years_salary
        self.years_count = years_count
        self.prof_salary = prof_salary
        self.prof_count = prof_count
        self.areas_salary = areas_salary
        self.areas_count = areas_count

    @classmethod
    def give_bar_subplot(cls, figure: figure, title: str, width: int, full_dict: dict,
                         x1_label: str, axis, subplot_type="",
                         prof_dict={}, x2_label=""):
        ax = figure.add_subplot(width)
        ax.set_title(title, fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis=axis)
        if subplot_type == "horizontal":
            ax.barh(list(full_dict.keys()), list(full_dict.values()), label=x1_label, align="center")
            ax.invert_yaxis()
        else:
            x_axis = range(len(full_dict.keys()))
            x1 = list(map(lambda x: float(x) - 0.2, x_axis))
            x2 = list(map(lambda x: float(x) + 0.2, x_axis))
            ax.bar(x1, list(full_dict.values()), width=0.4, label=x1_label)
            ax.bar(x2, list(prof_dict.values()), width=0.4, label=x2_label)
            ax.set_xticks(x_axis, list(full_dict.keys()), rotation="vertical")
        ax.legend(fontsize=8)

    def gain_pays_graph(self):
        figure = plt.figure()
        plt.rcParams.update({'font.size': 8})
        self.give_bar_subplot(figure, "Уровень зарплат по годам", 221,
                              self.years_salary, "средняя з/п", "y",
                              subplot_type="", prof_dict=self.prof_salary,
                              x2_label=f'з/п {input_set.profession}')
        self.give_bar_subplot(figure, "Количество вакансий по годам", 222,
                              self.years_count, "Количество вакансий", "y",
                              subplot_type="", prof_dict=self.prof_count,
                              x2_label=f'Количество вакансий {input_set.profession}')
        self.give_bar_subplot(figure, "Уровень зарплат по городам", 223,
                              self.areas_salary, "уровень з/п", "x",
                              subplot_type="horizontal")
        self.give_pie_sublot(figure, "Доля вакансий по городам", 224, self.areas_count)

        plt.tight_layout()
        plt.savefig("graph.png", dpi=300)

    @classmethod
    def give_pie_sublot(cls, fig: figure, title: str, width: int, data: dict):
        ax = fig.add_subplot(width)
        ax.set_title(title, fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        keys = list(data.keys())
        values = list(data.values())
        ax.pie(values, labels=keys)


input_set = Input()
years = GraphData(input_set.pars(), "years")
prof_years = GraphData(input_set.pars(), "years", input_set.profession)
areas = GraphData(input_set.pars(), "areas")


def generate_Excel():
    reportExcel = ExcelReport(Side(style="thin", color="000000"), Font(bold=True))
    reportExcel.exel([years.gain_graph()[0], years.gain_graph()[1],
                      prof_years.gain_graph()[0], prof_years.gain_graph()[1],
                      areas.gain_graph()[0], areas.gain_graph()[1]])


generate_Excel()


def generate_Png():
    stroke = [years.gain_graph()[0], years.gain_graph()[1],
              prof_years.gain_graph()[0], prof_years.gain_graph()[1],
              areas.gain_graph()[0], areas.gain_graph()[1]]
    png = PngReport(stroke[0], stroke[1], stroke[2], stroke[3], stroke[4], stroke[5])
    png.gain_pays_graph()


generate_Png()


def generate_pdf(self):
    env = Environment(loader=FileSystemLoader('../templates'))
    template = env.get_template("pdf_template.html")
    stats = []
    for year in self.stats1.keys():
        stats.append([year, self.stats1[year], self.stats2[year], self.stats3[year], self.stats4[year]])
    for key in self.stats6:
        self.stats6[key] = round(self.stats6[key] * 100, 2)
    pdf_template = template.render(
        {'name': self.vacancy_name, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(),
                                                             'graph.png'),
         'stats': stats, 'stats5': self.stats5, 'stats6': self.stats6})
    pdfkit.from_string(pdf_template, 'report.pdf', options={"enable-local-file-access": ""})


generate_pdf()

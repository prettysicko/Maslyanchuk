import csv
import math
from openpyxl.worksheet import worksheet
from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Font
from openpyxl.worksheet.dimensions import DimensionHolder, ColumnDimension
from openpyxl.utils import get_column_letter


class Salary:
    currency_to_rub = {"AZN": 35.68,
                       "BYR": 23.91,
                       "EUR": 59.90,
                       "GEL": 21.74,
                       "KGS": 0.76,
                       "KZT": 0.13,
                       "RUR": 1,
                       "UAH": 1.64,
                       "USD": 60.66,
                       "UZS": 0.0055}

    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency

    def turn_ruble(self) -> float:
        return self.turn()

    def turn(self):
        value = float(self.currency_to_rub[self.salary_currency])
        return ((float(self.salary_from) + float(self.salary_to)) / 2) * value


class Vacancy:
    def __init__(self, vacancy_dict):
        self.table_first_line(vacancy_dict)

    def table_first_line(self, vacancy_dict):
        self.dict = vacancy_dict
        self.salary = Salary(self.dict['salary_from'],
                             self.dict['salary_to'],
                             self.dict['salary_currency'])


class Input:
    def __init__(self):
        self.file_name = input("Введите название файла: ")
        self.profession = input("Введите название профессии: ")
        self.fields = []

    def parser_cod(self, all_vacancy, read):
        for row in read:
            new_vacancy = self.gain_vacancies(row)
            all_vacancy.append(new_vacancy)
        return all_vacancy

    def pars(self):
        read = self.read_cSV(self.file_name)
        all_vacancy = []
        return self.parser_cod(all_vacancy, read)

    def read_cSV(self, file_name) -> []:
        with open(file_name, encoding="utf-8-sig") as test:
            unpack = csv.reader(test)
            date = []
            c = 0
            for row in unpack:
                if c < len(row):
                    c = len(row)
                if '' not in row and c == len(row):
                    date.append(row)
            self.fields = date[0]
            return date[1:]

    def gain_vacancies(self, row: []) -> Vacancy:
        vacancy = Vacancy(dict(zip(self.fields, row)))
        return vacancy


class GraphData:
    def __init__(self, data, x_axis, profession="not"):
        self.data = data
        self.profession = profession
        self.salary_data = dict()
        self.count_data = dict()
        self.calc_date(x_axis)
        self.x_axis = x_axis

    def data_from_vacancies(self, vacancy: Vacancy, x_axis):
        if x_axis == "years":
            abs = int(vacancy.dict['published_at'].split('-')[0])
        else:
            abs = vacancy.dict['area_name']
        if abs not in self.salary_data:
            self.salary_data[abs] = 0
        if abs not in self.count_data:
            self.count_data[abs] = 0
        salary = vacancy.salary.turn_ruble()
        if self.profession != "not" and self.profession not in vacancy.dict['name']:
            return
        self.renew_direct(abs, salary)

    def calc_date(self, x_axis):
        for vacancy in self.data:
            self.data_from_vacancies(vacancy, x_axis)
        for x in self.salary_data:
            if self.count_data[x] != 0:
                self.salary_data[x] = math.floor(self.salary_data[x] / self.count_data[x])

    def renew_direct(self, key: str, value: float):
        self.renew(key, value)

    def renew(self, key, value):
        try:
            self.salary_data[key] += value
            self.count_data[key] += 1
        except:
            self.salary_data[key] = value
            self.count_data[key] = 1

    def gain_graph(self):
        first_printed_dict = self.salary_data
        second_printed_dict = self.count_data
        if self.x_axis == "areas":
            vac_count = sum(list(second_printed_dict.values()))
            first_printed_dict = self.sorted_dict(
                dict(list(filter(lambda x: self.count_data[x[0]] / vac_count > 0.01, self.salary_data.items()))))
            second_printed_dict = self.sorted_dict(
                dict(list(filter(lambda x: self.count_data[x[0]] / vac_count > 0.01, self.count_data.items()))))
            for x in second_printed_dict:
                second_printed_dict[x] = float("%.4f" % (second_printed_dict[x] / vac_count))
        return first_printed_dict, second_printed_dict

    @classmethod
    def sorted_dict(cls, non_sorted_dict: dict) -> dict:
        return dict(list(sorted(non_sorted_dict.items(), key=lambda x: x[1], reverse=True))[:10])


class ExcelReport:
    def __init__(self, side, font):
        self.border = Border(left=side, top=side, right=side, bottom=side)
        self.font = font

    def exel(self, direct: list):
        work_book = Workbook()
        years_sheet = self.make_years(work_book, direct[0], direct[1], direct[2], direct[3])
        areas_sheet = self.make_areas(work_book, direct[4], direct[5])
        ws = work_book.worksheets[0]
        work_book.remove(ws)
        work_book.save("report.xlsx")

    def make_years(self, work_book: Workbook, salaries: dict, prof_salaries: dict, counts: dict,
                   prof_counts: dict) -> worksheet.Worksheet:
        year_leaf = work_book.create_sheet("Статистика по годам")
        year_leaf['A1'] = 'Год'
        year_leaf['B1'] = 'Средняя зарплата'
        year_leaf['C1'] = f'Средняя зарплата - {input_set.profession}'
        year_leaf['D1'] = 'Количество вакансий'
        year_leaf['E1'] = f'Количество вакансий - {input_set.profession}'
        column_cells = [year_leaf["A1"], year_leaf['B1'], year_leaf['C1'], year_leaf['D1'], year_leaf['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        dim_holder = self.pays(counts, prof_counts, prof_salaries, salaries, year_leaf)

        for col in range(year_leaf.min_column, year_leaf.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(year_leaf, min=col, max=col, width=20)

        year_leaf.column_dimensions = dim_holder

        return year_leaf

    def pays(self, counts, prof_counts, prof_salaries, salaries, year_leaf):
        for x in range(len(salaries)):
            year_leaf[f'A{x + 2}'] = list(salaries.keys())[x]
            year_leaf[f'B{x + 2}'] = list(salaries.values())[x]
            year_leaf[f'C{x + 2}'] = list(prof_salaries.values())[x]
            year_leaf[f'D{x + 2}'] = list(counts.values())[x]
            year_leaf[f'E{x + 2}'] = list(prof_counts.values())[x]
            year_leaf[f'A{x + 2}'].border = self.border
            year_leaf[f'B{x + 2}'].border = self.border
            year_leaf[f'C{x + 2}'].border = self.border
            year_leaf[f'D{x + 2}'].border = self.border
            year_leaf[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=year_leaf)
        return dim_holder

    def make_areas(self, work_book: Workbook, salaries: dict, counts: dict) -> worksheet.Worksheet:
        area_leaf = work_book.create_sheet("Статистика по городам")
        area_leaf['A1'] = "Город"
        area_leaf['B1'] = "Уровень зарплат"
        area_leaf['D1'] = "Город"
        area_leaf['E1'] = "Доля вакансий"
        column_cells = [area_leaf['A1'], area_leaf['B1'], area_leaf['D1'], area_leaf['E1']]
        for cell in column_cells:
            cell.border = self.border
            cell.font = self.font
        dim_holder = self.place(area_leaf, counts, salaries)

        for col in range(area_leaf.min_column, area_leaf.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(area_leaf, min=col, max=col, width=20)

        area_leaf.column_dimensions = dim_holder
        return area_leaf

    def place(self, area_leaf, counts, salaries):
        for x in range(10):
            area_leaf[f'A{x + 2}'] = list(salaries.keys())[x]
            area_leaf[f'B{x + 2}'] = list(salaries.values())[x]
            area_leaf[f'D{x + 2}'] = list(counts.keys())[x]
            area_leaf[f'E{x + 2}'] = list(counts.values())[x]
            area_leaf[f'A{x + 2}'].border = self.border
            area_leaf[f'B{x + 2}'].border = self.border
            area_leaf[f'D{x + 2}'].border = self.border
            area_leaf[f'E{x + 2}'].border = self.border
        dim_holder = DimensionHolder(worksheet=area_leaf)
        return dim_holder


input_set = Input()
years = GraphData(input_set.pars(), "years")
prof_years = GraphData(input_set.pars(), "years", input_set.profession)
areas = GraphData(input_set.pars(), "areas")
reportExcel = ExcelReport(Side(style="thin", color="000000"), Font(bold=True))
reportExcel.exel([years.gain_graph()[0], years.gain_graph()[1],
                  prof_years.gain_graph()[0], prof_years.gain_graph()[1],
                  areas.gain_graph()[0], areas.gain_graph()[1]])

